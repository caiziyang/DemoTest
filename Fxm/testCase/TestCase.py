import openpyxl
import requests
from requests import Response


class TestCase():
    def test_getProduct(self):
        url= "https://mt.lhs11.com/fxm/services/local/fxm/product/product@query.japi"
        params = {"pagenumber": 1, "pagesize": 10}
        headers = {"cookie": "RSESSIONID=CZ3R3H65OCN3N6RSZAW4IDIJ3DMYE7OU", "Content-Type": "application/json"}
        r = requests.get(url, params=params, headers=headers).json()
        data = r["result"]["rows"]
        for i in range(10):
            print(data[i])
            print(data[i]["product_code"])
        return data

        ##exceldata = self.test_excel_data(data)


    def test_list(self, data: list):
        for i in range(len(data)):
            return data[i]


    def test_list_dict(self,data_list:dict):
        values = list(data_list.keys())
        values = set(values)
        values = list(values)
        return values

    def test_q(self):
        t = self.test_getProduct()
        f = self.test_list(t)
        for k, v in f.items():
            print(k + ":" + str(v))




    def test_w_data(self, exceldata :dict):
        ##wb = openpyxl.load_workbook(filename="testdata.xlsx")  ##读取Excel
        wb = openpyxl.Workbook()
        sheet = wb.active
        row = sheet.min_row
        print(row)
        for data_row in range(row, 2):
            for data_col1 in range(1, 10):
                ##针对每一行下面还要进行for循环来写入列的数据
                _ = sheet.cell(row=data_row, column=data_col1, value=str(exceldata))

            wb.save(filename="testdata.xlsx")  # 保存数据
            print("保存成功")



    def inspect_response(self,r: Response):
        print(r["result"]["rows"])












